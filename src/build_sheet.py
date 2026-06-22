"""Stage 4/5: merge scrape + enrichment, dedupe, score channel, export xlsx."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .common import DATA, norm_phone

COLUMNS = [
    "business_name", "channel", "owner_name", "phone", "email",
    "instagram", "facebook", "website", "address", "city",
    "category", "rating", "reviews", "google_maps_url", "date_added",
]
# Stable cross-run identity. Not shown in the sheet; stored in the ledger.
KEY_FIELD = "key"


def to_row(place: dict, contact: dict, dm_channels: list[str]) -> dict:
    row = {
        "business_name": place.get("title", ""),
        "owner_name": contact.get("owner_name", ""),
        "phone": place.get("phone", "") or place.get("phoneUnformatted", ""),
        "email": contact.get("email", ""),
        "instagram": contact.get("instagram", ""),
        "facebook": contact.get("facebook", ""),
        "website": place.get("website", ""),
        "address": place.get("address", "")
        or place.get("street", ""),
        "city": place.get("city", ""),
        "category": place.get("categoryName", "")
        or (place.get("categories") or [""])[0],
        "rating": place.get("totalScore", ""),
        "reviews": place.get("reviewsCount", ""),
        "google_maps_url": place.get("url", ""),
        "date_added": date.today().isoformat(),
    }
    row["channel"] = score_lead(row, dm_channels)
    # Stable identity: Google place id first, then phone, then name+address.
    pid = str(place.get("place_id") or place.get("placeId")
              or place.get("cid") or "").strip()
    row[KEY_FIELD] = (
        pid
        or norm_phone(row["phone"])
        or f"{row['business_name'].lower().strip()}|"
           f"{row['address'].lower().strip()}"
    )
    return row


def score_lead(row: dict, dm_channels: list[str]) -> str:
    """DM-first routing: any social -> DM queue, else email, else needs-research."""
    has_social = any(row.get(ch) for ch in dm_channels)
    if has_social:
        return "DM"
    if row.get("email"):
        return "Email"
    return "Needs research"


def _sort_key(row: dict):
    """Rank by social proof: rating x log-ish review weight."""
    try:
        rating = float(row.get("rating") or 0)
    except (TypeError, ValueError):
        rating = 0.0
    try:
        revs = int(row.get("reviews") or 0)
    except (TypeError, ValueError):
        revs = 0
    return -(rating * min(revs, 500))


def build_rows(places: list[dict], contacts: list[dict], cfg: dict) -> list[dict]:
    """Make scored rows for one scrape batch, deduped within the batch by key."""
    dm_channels = cfg["dm_channels"]
    rows = [to_row(p, c, dm_channels) for p, c in zip(places, contacts)]
    seen, out = set(), []
    for r in rows:
        k = r.get(KEY_FIELD)
        if k and k in seen:
            continue
        if k:
            seen.add(k)
        out.append(r)
    return out


def _write_tab(wb: Workbook, title: str, rows: list[dict]):
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for c, col in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    for r, row in enumerate(rows, 2):
        for c, col in enumerate(COLUMNS, 1):
            ws.cell(r, c, row.get(col, ""))
    widths = {"business_name": 28, "email": 26, "website": 30,
              "instagram": 26, "facebook": 26, "address": 32,
              "google_maps_url": 22}
    for c, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def write_workbook(rows: list[dict], cfg: dict, path=None) -> str:
    """Write the deliverable xlsx (DM First / Email Second / All) from rows."""
    rows = sorted(rows, key=_sort_key)
    dm = [r for r in rows if r.get("channel") == "DM"]
    email = [r for r in rows if r.get("channel") == "Email"]

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    _write_tab(wb, "DM First", dm)
    _write_tab(wb, "Email Second", email)
    _write_tab(wb, "All", rows)

    out = Path(path) if path else (DATA / cfg["output"]["xlsx_name"])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[sheet] {len(rows)} leads -> {out}")
    print(f"[sheet]   DM First: {len(dm)} | Email Second: {len(email)} | "
          f"Needs research: {len(rows) - len(dm) - len(email)}")
    return str(out)
